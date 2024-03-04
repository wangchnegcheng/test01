# -*- encoding=utf8 -*-
import pandas as pd
import json
from openpyxl import load_workbook
import openpyxl
import os
from openpyxl.styles import Font


class DataConverter:
    def __init__(self, json_data_file_path):
        self.json_data = json_data_file_path
        self.sheet_name_dict = {
            "impedance_df": "Impedance Measurement",
            "battery_df": "Battery Measurement",
            "storage_current_df": "Storage Current Measurement",
            "sensing_current_df": "Sensing Current Measurement",
            "ble_current_df": "BLE Current Measurement",
            "pacing_01ms_df": "0.1ms pacing width",
            "pacing_04ms_df": "0.4ms pacing width",
            "pacing_15ms_df": "1.5ms pacing width",
            "pacing_05v_df": "0.5V pacing Amplitude",
            "pacing_35v_df": "3.5V pacing Amplitude",
            "pacing_75v_df": "7.5V pacing Amplitude",
            "pacing_6vATP_df": "6.0V ATP Amplitude Width",
            "pacing_rate_30bpm_df": "pacing rate 30bpm",
            "pacing_rate_40bpm_df": "pacing rate 40bpm",
            "pacing_rate_150bpm_df": "pacing rate 150bpm",
            "sensitivity_015mv_df": "0.15mv Sensitivity Test",
            "sensitivity_03mv_df": "0.3mv Sensitivity Test",
            "sensitivity_12mv_df": "1.2mv Sensitivity Test",
            "input_impedance_100k_df": "100K input Impedance",
            "input_impedance_500k_df": "500K input Impedance",
            "input_impedance_1000k_df": "1000K input Impedance",
            "escape_interval_40bpm_df": "40bpm Escape Interval",
            "escape_interval_30bpm_df": "30bpm Escape Interval",
            "escape_interval_150bpm_df": "150bpm Escape Interval",
            "sense_refractory_120ms_df": "120ms Sense Refractory",
            "sense_refractory_100ms_df": "100ms Sense Refractory",
            "sense_refractory_250ms_df": "250ms Sense Refractory",
            "pace_refractory_150ms_df": "150ms Pace Refractory",
            "pace_refractory_250ms_df": "250ms Pace Refractory",
            "pace_refractory_500ms_df": "500ms Pace Refractory",
            "shock_01j_df": "0.1J Shock",
            "shock_20j_df": "20J Shock",
            "shock_40j_df": "40J Shock",
            "shock_40j_charge_time_df": "40J Charge Time",
            "shock_40j_dump_voltage_df": "40J Dump Voltage",
            "pacing_50hz_df": "50hz pacing",
            "info_storage_test_df": "Info Storage Test",
            "read_temperature_df": "Read temperature(Manual check)",
            "shock_40j_10times_df": "10times 40J Shock",
            "firmware_version_df": "Firmware Version",
        }


    def read_json_file(self, file_path):
            with open(file_path, 'r') as file:
                data = json.load(file)
            return data

    def convert_impedance_to_dataframe(self, measurement_type):
            data = self.read_json_file(self.json_data)
            near_field_impedance = []
            far_field_impedance = []
            tolerance_near_field = None
            tolerance_far_field = None

            if measurement_type in data:
                impedance_data = data[measurement_type]

                if 'Tolerance' in impedance_data:
                    tolerance_data = impedance_data['Tolerance']

                    # 检查 NearField 和 FarField 的 Tolerance
                    if 'NearField' in tolerance_data:
                        tolerance_near_field = tolerance_data['NearField']['Tolerance']
                        print(tolerance_near_field)
                    if 'FarField' in tolerance_data:
                        tolerance_far_field = tolerance_data['FarField']['Tolerance']
                        print(tolerance_far_field)
                for k, v in impedance_data.items():
                    if k == 'Tolerance':
                        continue

                    for k1, v1 in v.items():
                        if 'NearField' in k1:
                            near_field_impedance.append([k, k1, v1['Impedance'], v1['TimeStamp'], tolerance_near_field])
                        else:
                            far_field_impedance.append([k, k1, v1['Impedance'], v1['TimeStamp'], tolerance_far_field])

            if not near_field_impedance and not far_field_impedance:
                return pd.DataFrame()


            return pd.DataFrame(near_field_impedance + far_field_impedance,
                                columns=['Measurement Type', 'Impedance Type', 'Impedance', 'TimeStamp', 'Tolerance'])

    def convert_battery_measurement_to_dataframe(self,measurement_type):
            data = self.read_json_file(self.json_data)
            if measurement_type in data:
                battery_measurement = data[measurement_type]
                df = pd.DataFrame(columns=['Measurement Type', 'Battery Voltage', 'Time Stamp', 'Voltage Tolerance'])

                if 'Tolerance' in battery_measurement and 'Voltage' in battery_measurement['Tolerance']:
                    voltage_tolerance = battery_measurement['Tolerance']['Voltage']['Tolerance']

                    if 'Battery Voltage Measurement' in battery_measurement:
                        battery_voltage_data = battery_measurement['Battery Voltage Measurement']
                        for i, (battery_voltage_name, battery_voltage_values) in enumerate(
                                battery_voltage_data.items()):
                            battery_voltage = battery_voltage_values['BatteryVoltage']
                            time_stamp = battery_voltage_values['TimeStamp']
                            df.loc[i] = [battery_voltage_name, battery_voltage, time_stamp, voltage_tolerance]

                return df

            return pd.DataFrame()

    def convert_measurement_to_dataframe(self,measurement_type, key, column_name):
            data = self.read_json_file(self.json_data)
            if measurement_type in data:
                test_data = data[measurement_type]
                df = pd.DataFrame(columns=['Measurement Type', column_name, 'Tolerance'])
                if 'Tolerance' in test_data and key in test_data['Tolerance']:
                    tolerance = test_data['Tolerance'][key]['Tolerance']
                    df.loc[0] = [measurement_type, test_data[column_name], tolerance]
                return df
            return pd.DataFrame()

    def convert_pace_to_dataframe(self, measurement_type, *tolerance_keys):
            data = self.read_json_file(self.json_data)
            if measurement_type not in data:
                # print(f"No data found for pacing width: {measurement_type}")
                return pd.DataFrame()

            test_data = data[measurement_type]

            if len(tolerance_keys) == 1:
                tolerance_key = tolerance_keys[0]
                df = pd.DataFrame(columns=['Measurement Type', 'Pace Result Detail', 'Amplitude', 'Width', 'Interval',
                                           'Averaged Amplitude', 'Rate', f"{tolerance_key} Tolerance"])

                if 'Tolerance' in test_data and tolerance_key in test_data['Tolerance']:
                    tolerance = test_data['Tolerance'][tolerance_key]['Tolerance']
                else:
                    print("Invalid tolerance key provided")
                    print(f"Available tolerance keys: {list(test_data.get('Tolerance', {}).keys())}")
                    return pd.DataFrame()
            elif len(tolerance_keys) == 2:
                tolerance_key1, tolerance_key2 = tolerance_keys
                if 'Tolerance' in test_data and tolerance_key1 in test_data['Tolerance'] and tolerance_key2 in \
                        test_data['Tolerance']:
                    tolerance1 = test_data['Tolerance'][tolerance_key1]['Tolerance']
                    tolerance2 = test_data['Tolerance'][tolerance_key2]['Tolerance']
                else:
                    print("Invalid tolerance keys provided")
                    print(
                        f"Available tolerance keys: {list(test_data.get('Tolerance', {}).get(tolerance_key1, {}).keys())}")
                    return pd.DataFrame()

                df = pd.DataFrame(columns=['Measurement Type', 'Pace Result Detail', 'Amplitude', 'Width', 'Interval',
                                           'Averaged Amplitude', 'Rate', f"{tolerance_key1} Tolerance",
                                           f"{tolerance_key2} Tolerance"])
            elif len(tolerance_keys) == 3:
                tolerance_key1, tolerance_key2, tolerance_key3 = tolerance_keys
                if 'Tolerance' in test_data and tolerance_key1 in test_data['Tolerance'] and tolerance_key2 in \
                        test_data['Tolerance'] and tolerance_key3 in test_data['Tolerance']:
                    tolerance1 = test_data['Tolerance'][tolerance_key1]['Tolerance']
                    tolerance2 = test_data['Tolerance'][tolerance_key2]['Tolerance']
                    tolerance3 = test_data['Tolerance'][tolerance_key3]['Tolerance']
                else:
                    print("Invalid tolerance keys provided")
                    print(
                        f"Available tolerance keys: {list(test_data.get('Tolerance', {}).get(tolerance_key1, {}).keys())}")
                    return pd.DataFrame()

                df = pd.DataFrame(columns=['Measurement Type', 'Pace Result Detail', 'Amplitude', 'Width', 'Interval',
                                           'Averaged Amplitude', 'Rate', f"{tolerance_key1} Tolerance",
                                           f"{tolerance_key2} Tolerance", f"{tolerance_key3} Tolerance"])

            else:
                print("Invalid number of tolerance keys provided")
                return pd.DataFrame()

            if isinstance(test_data, dict):
                for pace_name, pace_result_detail in test_data.items():
                    if pace_name != 'Tolerance':
                        if isinstance(pace_result_detail, dict):
                            for detail_name, pace_result in pace_result_detail.items():
                                amplitude = pace_result.get('Amplitude', None)
                                width = pace_result.get('Width', None)
                                interval = pace_result.get('Interval', None)
                                averaged_amplitude = pace_result.get('AveragedAmplitude', None)
                                rate = pace_result.get('Rate', None)

                                if len(tolerance_keys) == 1:
                                    df.loc[len(df)] = [measurement_type, f"{pace_name}-{detail_name}", amplitude, width,
                                                       interval, averaged_amplitude, rate, tolerance]
                                elif len(tolerance_keys) == 2:
                                    df.loc[len(df)] = [measurement_type, f"{pace_name}-{detail_name}", amplitude, width,
                                                       interval, averaged_amplitude, rate, tolerance1, tolerance2]
                                else:
                                    df.loc[len(df)] = [measurement_type, f"{pace_name}-{detail_name}", amplitude, width,
                                                       interval, averaged_amplitude, rate, tolerance1, tolerance2,
                                                       tolerance3]
                        else:
                            print(f"Invalid data format: {pace_result_detail}")

            return df

    def convert_sensitivity_test_to_dataframe(self,measurement_type):
            data = self.read_json_file(self.json_data)
            if measurement_type in data:
                test_data = data[measurement_type]
                df = pd.DataFrame(columns=['Measurement Type', 'MinAmplitude', 'ErrorRate', 'Sensitivity Tolerance'])
                if 'Tolerance' in test_data and 'Sensitivity' in test_data['Tolerance']:
                    sensitivity_tolerance = test_data['Tolerance']['Sensitivity']['Tolerance']
                    df.loc[0] = [measurement_type, test_data['MinAmplitude'], test_data['ErrorRate'],
                                 sensitivity_tolerance]
                return df

            return pd.DataFrame()

    def convert_input_impedance_dataframe(self,measurement_type, input_range):
            data = self.read_json_file(self.json_data)
            if measurement_type in data:
                test_data = data[measurement_type]
                # print(test_data)
                df = pd.DataFrame(columns=['Measurement Type', f'{input_range} input impedance', 'Tolerance'])
                if 'Tolerance' in test_data and 'InputImpedance' in test_data['Tolerance']:
                    inputImpedance_tolerance = test_data['Tolerance']['InputImpedance']['Tolerance']
                    df.loc[0] = [measurement_type, test_data[f'{input_range} input impedance'],
                                 inputImpedance_tolerance]
                return df

            return pd.DataFrame()

    def convert_shock_to_dataframe(self, measurement_type):
            data = self.read_json_file(self.json_data)
            rows = []
            if measurement_type not in data:
                # print(f"Measurement type '{measurement_type}' not found in data")
                return pd.DataFrame()
            if data:

                if 'Tolerance' in data[measurement_type]:
                    tolerance_data = data[measurement_type]['Tolerance']
                    if 'ChargeVoltage' in tolerance_data and 'ShockVoltage' in tolerance_data and 'DeliveredEnergy' in tolerance_data:
                        tolerance1 = tolerance_data['ChargeVoltage']['Tolerance']
                        tolerance2 = tolerance_data['ShockVoltage']['Tolerance']
                        tolerance3 = tolerance_data['DeliveredEnergy']['Tolerance']
                    else:
                        print("Invalid tolerance keys provided")
                        return pd.DataFrame()
                else:
                    print(f"No tolerance data found for measurement type: {measurement_type}")
                    return pd.DataFrame()

                for key, pacing_data in data.items():
                    if key == measurement_type:  # 只处理符合特定measurement_type的数据
                        if 'Charge Voltage' not in pacing_data or 'EstimatedImpedance' not in pacing_data:
                            continue
                        charge_voltage = pacing_data['Charge Voltage']
                        est_impedance = pacing_data['EstimatedImpedance']
                        shock_result_detail = pacing_data['ShockDetail']['ShockResultDetail1']
                        phase1_start = shock_result_detail['Phase1Start']
                        phase1_end = shock_result_detail['Phase1End']
                        phase2_start = shock_result_detail['Phase2Start']
                        phase2_end = shock_result_detail['Phase2End']
                        phase1_percent = shock_result_detail['Phase1Percent']
                        phase2_percent = shock_result_detail['Phase2Percent']
                        difference_between_two_phase = shock_result_detail['DifferenceBetweenTwoPhase']
                        phase1_interval = shock_result_detail['Phase1Interval']
                        phase2_interval = shock_result_detail['Phase2Interval']
                        between_phase_interval = shock_result_detail['BetweenPhaseInterval']
                        energy = shock_result_detail['Energy']
                        calculated_energy = shock_result_detail['Calculated Energy']
                        directon = shock_result_detail['Directoon']


                        rows.append(
                            [measurement_type, charge_voltage, calculated_energy , est_impedance, phase1_start,
                             phase1_end, phase2_start,phase2_end, phase1_percent, phase2_percent,
                             difference_between_two_phase, phase1_interval,phase2_interval, between_phase_interval,
                             energy, directon, tolerance1,tolerance2, tolerance3])

            shock_df = pd.DataFrame(rows, columns=['Measurement Type', 'Charge Voltage', 'Calculated Energy',
                                                   'Estimated Impedance','Phase 1 Start', 'Phase 1 End',
                                                   'Phase 2 Start', 'Phase 2 End', 'Phase 1 Percent', 'Phase 2 Percent',
                                                   'Difference Between Two Phases','Phase 1 Interval',
                                                   'Phase 2 Interval', 'Between Phase Interval',
                                                   'Energy', 'Directoon',
                                                   'chargeVoltage tolerance', 'shockVoltage tolerance',
                                                   'deliveredEnergy tolerance'])

            return shock_df

    def convert_shock_10_times_to_dataframe(self, measurement_type):
        rows = []
        data = self.read_json_file(self.json_data)

        if measurement_type not in data:
            # print(f"Measurement type '{measurement_type}' not found in data")
            return pd.DataFrame()

        shock_data = data[measurement_type]
        tolerance_data = shock_data["Tolerance"]

        if "ChargeVoltage" not in tolerance_data or "ShockVoltage" not in tolerance_data or "DeliveredEnergy" not in tolerance_data:
            print("Invalid tolerance keys provided")
            return pd.DataFrame()

        tolerance1 = tolerance_data["ChargeVoltage"]["Tolerance"]
        tolerance2 = tolerance_data["ShockVoltage"]["Tolerance"]
        tolerance3 = tolerance_data['DeliveredEnergy']['Tolerance']

        for shock_num, shock_detail in shock_data["ShockDetail"].items():
            charge_voltage = shock_detail["Charge Result"]
            phase1_start = shock_detail["Shock Detail"]["Phase1Start"]
            phase1_end = shock_detail["Shock Detail"]["Phase1End"]
            phase2_start = shock_detail["Shock Detail"]["Phase2Start"]
            phase2_end = shock_detail["Shock Detail"]["Phase2End"]
            phase1_percent = shock_detail["Shock Detail"]["Phase1Percent"]
            phase2_percent = shock_detail["Shock Detail"]["Phase2Percent"]
            difference_between_two_phase = shock_detail["Shock Detail"]["DifferenceBetweenTwoPhase"]
            phase1_interval = shock_detail["Shock Detail"]["Phase1Interval"]
            phase2_interval = shock_detail["Shock Detail"]["Phase2Interval"]
            between_phase_interval = shock_detail["Shock Detail"]["BetweenPhaseInterval"]
            energy = shock_detail["Shock Detail"]['Energy']
            calculated_energy = shock_detail["Shock Detail"]['Calculated Energy']
            directon = shock_detail["Shock Detail"]['Directoon']

            rows.append([
                measurement_type,
                charge_voltage,
                calculated_energy,
                phase1_start,
                phase1_end,
                phase2_start,
                phase2_end,
                phase1_percent,
                phase2_percent,
                difference_between_two_phase,
                phase1_interval,
                phase2_interval,
                between_phase_interval,
                energy,
                directon,
                tolerance1,
                tolerance2,
                tolerance3,
            ])

        shock_df = pd.DataFrame(rows, columns=[
            "Measurement Type",
            "Charge Voltage",
            'Calculated Energy',
            "Phase 1 Start",
            "Phase 1 End",
            "Phase 2 Start",
            "Phase 2 End",
            "Phase 1 Percent",
            "Phase 2 Percent",
            "Difference Between Two Phases",
            "Phase 1 Interval",
            "Phase 2 Interval",
            "Between Phase Interval",
            'Energy',
            'Directoon',
            "chargeVoltage tolerance",
            "shockVoltage tolerance",
            "deliveredEnergy tolerance",
        ])

        return shock_df


    def convert_info_storage_test(self, measurement_type):
            data = self.read_json_file(self.json_data)
            if measurement_type in data:
                test_data = data[measurement_type]
                df = pd.DataFrame(columns=['MeasurementType', 'PatientInfo', 'ClinicianNote', 'LeadInfo'])
                df.loc[0] = [measurement_type, test_data.get('PatientInfo', ''), test_data.get('ClinicianNote', ''),
                             test_data.get('LeadInfo', '')]
                return df
            else:
                return pd.DataFrame()

    def convert_read_temperature(self, measurement_type):
            data = self.read_json_file(self.json_data)
            if measurement_type in data:
                temperature_measurement = data[measurement_type]
                df = pd.DataFrame(columns=['Measurement Type', 'Temperature', 'TimeStamp', 'Temperature Tolerance'])

                if 'Tolerance' in temperature_measurement and 'temperature' in temperature_measurement['Tolerance']:
                    temperature_tolerance = temperature_measurement['Tolerance']['temperature']['Tolerance']

                    if 'Battery Voltage Measurement' in temperature_measurement:
                        temperature_data = temperature_measurement['Battery Voltage Measurement']
                        for i, (temperature_name, temperature_info) in enumerate(temperature_data.items()):
                            if 'temperature' in temperature_info and 'TimeStamp' in temperature_info:
                                temperature = temperature_info['temperature']
                                time_stamp = temperature_info['TimeStamp']
                                df.loc[i] = [temperature_name, temperature, time_stamp, temperature_tolerance]

                return df

            return pd.DataFrame()

    def convert_firmware_version(self, measurement_type):
        data = self.read_json_file(self.json_data)
        if measurement_type in data:
            firmware_version = data[measurement_type]
            df = pd.DataFrame(columns=['Measurement Type', 'SoftwareRelease', 'CRC32InUse', 'M1_BuildNumber',
                                       'M2_BuildNumber', 'M3_BuildNumber', 'BLE_BuildNumber', 'HW_Ver'])

            df.loc[len(df)] = [measurement_type, firmware_version['SoftwareRelease'], firmware_version['CRC32InUse'],
                               firmware_version['M1_BuildNumber'], firmware_version['M2_BuildNumber'],
                               firmware_version['M3_BuildNumber'], firmware_version['BLE_BuildNumber'],
                               firmware_version['HW_Ver']]
            return df

        return pd.DataFrame()

    def convert_to_excel(self, output_file_path):

        # data = self.read_json_file(self.json_data)
        impedance_df = self.convert_impedance_to_dataframe(self.sheet_name_dict["impedance_df"])
        battery_df = self.convert_battery_measurement_to_dataframe(self.sheet_name_dict["battery_df"])
        storage_current_df = self.convert_measurement_to_dataframe(self.sheet_name_dict["storage_current_df"], 'MaxCurrent',
                                                                   'StorageModeCurrentUsage')
        sensing_current_df = self.convert_measurement_to_dataframe(self.sheet_name_dict["sensing_current_df"], 'MaxCurrent',
                                                                   '100%SensingCurrentUsage')
        ble_current_df = self.convert_measurement_to_dataframe(self.sheet_name_dict["ble_current_df"], 'MaxCurrent',
                                                               'BLESessionCurrentUsage')

        pacing_01ms_df = self.convert_pace_to_dataframe(self.sheet_name_dict["pacing_01ms_df"], 'Width')
        pacing_04ms_df = self.convert_pace_to_dataframe(self.sheet_name_dict["pacing_04ms_df"], 'Width')
        pacing_15ms_df = self.convert_pace_to_dataframe(self.sheet_name_dict["pacing_15ms_df"], 'Width')

        pacing_05v_df = self.convert_pace_to_dataframe(self.sheet_name_dict["pacing_05v_df"], 'Amplitude')
        pacing_35v_df = self.convert_pace_to_dataframe(self.sheet_name_dict["pacing_35v_df"], 'Amplitude')
        pacing_75v_df = self.convert_pace_to_dataframe(self.sheet_name_dict["pacing_75v_df"], 'Amplitude')

        pacing_6vATP_df = self.convert_pace_to_dataframe(self.sheet_name_dict["pacing_6vATP_df"], 'Width', 'Amplitude')
        pacing_rate_30bpm_df = self.convert_pace_to_dataframe(self.sheet_name_dict["pacing_rate_30bpm_df"], 'Interval')
        pacing_rate_40bpm_df = self.convert_pace_to_dataframe(self.sheet_name_dict["pacing_rate_40bpm_df"], 'Interval')
        pacing_rate_150bpm_df = self.convert_pace_to_dataframe(self.sheet_name_dict["pacing_rate_150bpm_df"], 'Interval')

        sensitivity_015mv_df = self.convert_sensitivity_test_to_dataframe(self.sheet_name_dict["sensitivity_015mv_df"])
        sensitivity_03mv_df = self.convert_sensitivity_test_to_dataframe(self.sheet_name_dict["sensitivity_03mv_df"])
        sensitivity_12mv_df = self.convert_sensitivity_test_to_dataframe(self.sheet_name_dict["sensitivity_12mv_df"])

        input_impedance_100k_df = self.convert_input_impedance_dataframe(self.sheet_name_dict["input_impedance_100k_df"], "100K")
        input_impedance_500k_df = self.convert_input_impedance_dataframe(self.sheet_name_dict["input_impedance_500k_df"], "500K")
        input_impedance_1000k_df = self.convert_input_impedance_dataframe(self.sheet_name_dict["input_impedance_1000k_df"], "1000K")

        escape_interval_40bpm_df = self.convert_measurement_to_dataframe(self.sheet_name_dict["escape_interval_40bpm_df"], 'Interval',
                                                                         'Escape Interval')
        escape_interval_30bpm_df = self.convert_measurement_to_dataframe(self.sheet_name_dict["escape_interval_30bpm_df"], 'Interval',
                                                                         'Escape Interval')
        escape_interval_150bpm_df = self.convert_measurement_to_dataframe(self.sheet_name_dict["escape_interval_150bpm_df"], 'Interval',
                                                                          'Escape Interval')

        sense_refractory_120ms_df = self.convert_measurement_to_dataframe(self.sheet_name_dict["sense_refractory_120ms_df"], 'Interval',
                                                                          'Sense Refractory')
        sense_refractory_100ms_df = self.convert_measurement_to_dataframe(self.sheet_name_dict["sense_refractory_100ms_df"], 'Interval',
                                                                          'Sense Refractory')
        sense_refractory_250ms_df = self.convert_measurement_to_dataframe(self.sheet_name_dict["sense_refractory_250ms_df"], 'Interval',
                                                                          'Sense Refractory')

        pace_refractory_150ms_df = self.convert_measurement_to_dataframe(self.sheet_name_dict["pace_refractory_150ms_df"], 'Interval',
                                                                         'Pace Refractory')
        pace_refractory_250ms_df = self.convert_measurement_to_dataframe(self.sheet_name_dict["pace_refractory_250ms_df"], 'Interval',
                                                                         'Pace Refractory')
        pace_refractory_500ms_df = self.convert_measurement_to_dataframe(self.sheet_name_dict["pace_refractory_500ms_df"], 'Interval',
                                                                         'Pace Refractory')

        shock_01j_df = self.convert_shock_to_dataframe(self.sheet_name_dict["shock_01j_df"])
        shock_20j_df = self.convert_shock_to_dataframe(self.sheet_name_dict["shock_20j_df"])
        shock_40j_df = self.convert_shock_to_dataframe(self.sheet_name_dict["shock_40j_df"])

        shock_40j_10times_df = self.convert_shock_10_times_to_dataframe(self.sheet_name_dict["shock_40j_10times_df"])

        shock_40j_charge_time_df = self.convert_measurement_to_dataframe(self.sheet_name_dict["shock_40j_charge_time_df"], 'Time',
                                                                         'Charge Duration')
        shock_40j_dump_voltage_df = self.convert_measurement_to_dataframe(self.sheet_name_dict["shock_40j_dump_voltage_df"], 'Voltage',
                                                                          'AfterDumpVoltage')
        pacing_50hz_df = self.convert_pace_to_dataframe(self.sheet_name_dict["pacing_50hz_df"], 'Width', 'Amplitude', 'Interval')
        info_storage_test_df = self.convert_info_storage_test(self.sheet_name_dict["info_storage_test_df"])

        read_temperature_df = self.convert_read_temperature(self.sheet_name_dict["read_temperature_df"])

        firmware_version_df = self.convert_firmware_version(self.sheet_name_dict["firmware_version_df"])


        with pd.ExcelWriter(output_file_path) as writer:
            for var_name, sheet_name in self.sheet_name_dict.items():
                df = locals().get(var_name)
                if df is not None and not df.empty:
                    df.to_excel(writer, sheet_name=sheet_name, index=False)



        wb = openpyxl.load_workbook(output_file_path)
        for sheet in wb.worksheets:
            for column in sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    if cell.value is not None:
                        cell_value = str(cell.value)
                        if len(cell_value) > max_length:
                            max_length = len(cell_value)
                adjusted_width = (max_length + 2) * 1.2
                sheet.column_dimensions[column_letter].width = adjusted_width
        wb.save(output_file_path)

    def mark_fail_projects_in_excel(self, output_file_path, json_results_path):
        data = self.read_json_file(json_results_path)
        fail_projects = [project for project, status in data.items() if status == "fail"]
        pass_projects = [project for project, status in data.items() if status == "pass"]

        workbook = load_workbook(output_file_path)

        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            if sheet_name in fail_projects:
                worksheet.sheet_properties.tabColor = "FF0000"
            if sheet_name in pass_projects:
                worksheet.sheet_properties.tabColor = "00FF00"

            if sheet_name == "50hz pacing":
                # 遍历所有的列，找到标题为"Rate"的列
                rate_column = None
                for cell in worksheet[1]:
                    if cell.value == "Rate":
                        rate_column = cell.column

                if rate_column is not None:
                    # 在"Rate"列后新增一列
                    new_column = rate_column + 1
                    worksheet.insert_cols(new_column)
                    # worksheet.cell(row=1, column=new_column, value="50HZ").font = Font(bold=True)  # 新列标题设置为"50HZ"并加粗显示
                    title_cell = worksheet.cell(row=1, column=new_column, value="50HZ")
                    title_cell.font = Font(bold=True)

                    for row in worksheet.iter_rows(min_row=2, min_col=rate_column, max_col=rate_column):
                        for cell in row:
                            try:
                                # 去除单元格中的"bpm"并转换为数字
                                cell_value = float(cell.value.replace("bpm", ""))
                                # 在新增列中写入转换后的数据（除以60并保留一位小数，再加上单位"Hz"）
                                worksheet.cell(row=cell.row, column=new_column, value=f"{round(cell_value / 60, 1)}hz")

                            except (ValueError, AttributeError):
                                # 如果转换失败（即单元格不是数字），则保留原值或处理错误
                                print(f"Cannot convert value '{cell.value}' in cell {cell.coordinate} to a float.")

        workbook.save(output_file_path)

def batch_convert_to_excel(folder_path):
    for root, dirs, files in os.walk(folder_path):
        data_file_path = None
        results_file_path = None

        for file in files:
            if file.endswith("test_data.json"):
                data_file_path = os.path.join(root, file)
            elif file.endswith("test_results.json"):
                results_file_path = os.path.join(root, file)

        # 检查是否同时存在匹配的测试数据和测试结果文件
        if data_file_path is not None and results_file_path is not None:
            # 创建 DataConverter 实例并传递文件路径
            dir_name = os.path.basename(os.path.dirname(root))
            sub_dir_name = os.path.basename(root)

            # 拼接为 Excel 文件名
            excel_file_name = f"{dir_name}_{sub_dir_name}.xlsx"
            output_file_path = os.path.join(root, excel_file_name)
            converter = DataConverter(data_file_path)
            # output_file_path = os.path.join(root, "output.xlsx")
            converter.convert_to_excel(output_file_path)
            converter.mark_fail_projects_in_excel(output_file_path, results_file_path)







# folder_path = r"C:\Users\wangchengcheng\Desktop\devices\ICD1022-E6_9D_DB_C8_11_FE\2024-02-26_15-21-01"
# batch_convert_to_excel(folder_path)
